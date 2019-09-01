from bs4 import BeautifulSoup as soup
from openpyxl import Workbook, load_workbook
from urllib.request import urlopen, Request
import random, os

user_agents = [
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)']


# This function creates a file to store the information.
def create_data_file():
    file_headers = ['Used or New', 'Year', 'Make', 'Model', 'Price', 'Milage', 'Drive Type',
                    'Engine', 'Transmission', 'Fuel Type', 'MPG', 'Interior', 'Exterior', 'VIN', 'Seller']
    wb = Workbook()
    ws = wb.active
    ws.append(file_headers)
    wb.save(filename='ProductData.xlsx')


# This function is for storing information.
def save_data(data):
    wb = load_workbook('ProductData.xlsx')
    ws = wb.active
    ws.append(data)
    wb.save('ProductData.xlsx')


# This function requests a web page.
def make_requests(url):
    headers = {"User-Agent": random.choice(user_agents)}
    page_html = urlopen(Request(url, headers=headers))
    return soup(page_html, "html.parser")


# This function extracts the link from all car ads.
def find_all_car_links(soup):
    print('\nCreating a list of ads for all cars ... \n')
    base_url = 'https://www.autotrader.com'
    link_list = []
    try:
        for i in soup.find_all('div', class_='display-flex justify-content-between')[1:]:
            href = i.a['href']
            link_list.append('{0}{1}'.format(base_url, href))
        return link_list
    except:
        print('Can not listing links')


# This function extracts information from each car individually.
def extract_page(soup):
    price = soup.find(
        'div', class_='text-gray-base text-bold text-size-600 margin-right-auto').text
    Name = soup.find(
        'h1', class_='text-bold text-size-600 text-size-sm-700 margin-right-2').text
    seller = soup.find(
        'div', class_='colored-background bg-blue-lightest').div.div.div.div.text.replace('Call DealerChat with Dealer', '')
    N = Name.split()
    new_or_used = N[0]
    year = N[1]
    make = N[2]
    model = str(N[3:])
    car_info = [i.div.find(class_='col-xs-8').text for i in soup.find_all(
        'li', class_='list-bordered list-condensed')]
    milage = car_info[0].replace(',', '')
    drive_type = car_info[1]
    engine = car_info[2]
    transmission = car_info[3]
    fuel_type = car_info[4]
    mpg = car_info[5]
    interior = car_info[6]
    exterior = car_info[7]
    stock_number = car_info[8]
    vin = car_info[9]

    all_data = [new_or_used, year, make, model, price, milage, drive_type,
                engine, transmission, fuel_type, mpg, interior, exterior, vin, seller]
    print(all_data)
    save_data(all_data)


def main():
    cars_count = 0

    if os.path.exists('/ProductData.xlsx') != True:
        create_data_file()

    url = input('\nInsert your link please: ')

    bs = make_requests(url)
    cars_links = find_all_car_links(bs)

    for car_link in cars_links:
        bs = make_requests(car_link)
        print('Car Number ' + str(cars_count) + ' -> Information:')
        cars_count += 1
        extract_page(bs)
        print('\n\n')


if __name__ == '__main__':
    main()
